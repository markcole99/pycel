import calendar
import collections
import datetime as dt
import operator
import re

from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import (
    get_column_letter,
    range_boundaries as openpyxl_range_boundaries
)


ERROR_CODES = frozenset(Tokenizer.ERROR_CODES)
DIV0 = '#DIV/0!'
VALUE_ERROR = '#VALUE!'

R1C1_ROW_RE_STR = r"R(\[-?\d+\]|\d+)?"
R1C1_COL_RE_STR = r"C(\[-?\d+\]|\d+)?"
R1C1_COORD_RE_STR = "(?P<row>{0})?(?P<col>{1})?".format(
    R1C1_ROW_RE_STR, R1C1_COL_RE_STR)

R1C1_COORDINATE_RE = re.compile('^' + R1C1_COORD_RE_STR + '$', re.VERBOSE)

R1C1_RANGE_EXPR = """
(?P<min_row>{0})?
(?P<min_col>{1})?
(:(?P<max_row>{0})?
(?P<max_col>{1})?)?
""".format(R1C1_ROW_RE_STR, R1C1_COL_RE_STR)

R1C1_RANGE_RE = re.compile('^' + R1C1_RANGE_EXPR + '$', re.VERBOSE)

TABLE_REF_RE = re.compile(r"^(?P<table_name>[^[]+)\[(?P<table_selector>.*)\]$")

TABLE_SELECTOR_RE = re.compile(
    r"^(?P<row_or_column>[^[]+)$|"
    r"^@\[(?P<this_row_column>[^[]*)\]$|"
    r"^ *(?P<rows>(\[([^\]]+)\] *, *)*)"
    r"(\[(?P<start_col>[^\]]+)\] *: *)?"
    r"(\[(?P<end_col>.+)\] *)?$")


MAX_COL = 16384
MAX_ROW = 1048576

VALID_R1C1_RANGE_ITEM_COMBOS = {
    (0, 1, 0, 1),
    (1, 0, 1, 0),
    (1, 1, 1, 1),
}

OPERATORS = {
    '': operator.eq,
    '<': operator.lt,
    '>': operator.gt,
    '<=': operator.le,
    '>=': operator.ge,
    '<>': operator.ne,
}

PYTHON_AST_OPERATORS = {
    'Eq': operator.eq,
    'Lt': operator.lt,
    'Gt': operator.gt,
    'LtE': operator.le,
    'GtE': operator.ge,
    'NotEq': operator.ne,
    'Add': operator.add,
    'Sub': operator.sub,
    'Mult': operator.mul,
    'Div': operator.truediv,
    'FloorDiv': operator.floordiv,
    'Mod': operator.mod,
    'Pow': operator.pow,
    'LShift': operator.lshift,
    'RShift': operator.rshift,
    'BitOr': operator.or_,
    'BitXor': operator.xor,
    'BitAnd': operator.and_,
    'MatMult': operator.matmul,
}


class PyCelException(Exception):
    """Base class for PyCel errors"""


class AddressRange(collections.namedtuple(
        'Address', 'address sheet start end coordinate')):

    def __new__(cls, address, *args, sheet=''):
        if args:
            return super(AddressRange, cls).__new__(cls, address, *args)

        if isinstance(address, str):
            return cls.create(address, sheet=sheet)

        elif isinstance(address, AddressCell):
            return AddressCell(address, sheet=sheet)

        elif isinstance(address, AddressRange):
            if not sheet or sheet == address.sheet:
                return address

            elif not address.sheet:
                start = AddressCell(address.start.coordinate, sheet=sheet)
                end = AddressCell(address.end.coordinate, sheet=sheet)

            else:
                raise ValueError("Mismatched sheets '{}' and '{}'".format(
                    address, sheet))

        else:
            assert (isinstance(address, tuple) and
                    4 == len(address) and
                    None in address or address[0:2] != address[2:]), \
                "AddressRange expected a range '{}'".format(address)

            start_col, start_row, end_col, end_row = address
            start = AddressCell(
                (start_col, start_row, start_col, start_row), sheet=sheet)
            end = AddressCell(
                (end_col, end_row, end_col, end_row), sheet=sheet)

        coordinate = '{0}:{1}'.format(start.coordinate, end.coordinate)

        format_str = '{0}!{1}' if sheet else '{1}'
        return super(AddressRange, cls).__new__(
            cls, format_str.format(sheet, coordinate),
            sheet, start, end, coordinate)

    def __str__(self):
        return self.address

    @property
    def is_range(self):
        return True

    @property
    def size(self):
        if 0 in (self.end.row, self.start.row):
            height = MAX_ROW
        else:
            height = self.end.row - self.start.row + 1

        if 0 in (self.end.col_idx, self.start.col_idx):
            width = MAX_COL
        else:
            width = self.end.col_idx - self.start.col_idx + 1

        return AddressSize(height, width)

    @property
    def has_sheet(self):
        return bool(self.sheet)

    @property
    def sort_key(self):
        return self.sheet, self.start.col_idx, self.start.row

    @property
    def rows(self):
        """Get each addresses for every cell, yields one row at a time."""
        col_range = self.start.col_idx, self.end.col_idx + 1
        for row in range(self.start.row, self.end.row + 1):
            yield (AddressCell((col, row, col, row), sheet=self.sheet)
                   for col in range(*col_range))

    @property
    def cols(self):
        """Get each addresses for every cell, yields one column at a time."""
        col_range = self.start.col_idx, self.end.col_idx + 1
        for col in range(*col_range):
            yield (AddressCell((col, row, col, row), sheet=self.sheet)
                   for row in range(self.start.row, self.end.row + 1))

    @classmethod
    def create(cls, address, sheet='', cell=None):
        if isinstance(address, AddressRange):
            return AddressRange(address, sheet=sheet)

        elif isinstance(address, AddressCell):
            return AddressCell(address, sheet=sheet)

        sheetname, addr = split_sheetname(address, sheet=sheet)
        addr_tuple, sheetname = range_boundaries(
            addr, sheet=sheetname, cell=cell)

        if None in addr_tuple or addr_tuple[0:2] != addr_tuple[2:]:
            return AddressRange(addr_tuple, sheet=sheetname)
        else:
            return AddressCell(addr_tuple, sheet=sheetname)


class AddressCell(collections.namedtuple(
        'AddressCell', 'address sheet col_idx row coordinate')):

    def __new__(cls, address, *args, sheet=''):
        if args:
            return super(AddressCell, cls).__new__(cls, address, *args)

        if isinstance(address, str):
            return cls.create(address, sheet=sheet)

        elif isinstance(address, AddressCell):
            if not sheet or sheet == address.sheet:
                return address

            elif not address.sheet:
                row, col_idx, coordinate = address[2:5]

            else:
                raise ValueError("Mismatched sheets '{}' and '{}'".format(
                    address, sheet))

        else:
            assert (isinstance(address, tuple) and
                    4 == len(address) and
                    None not in address or address[0:2] == address[2:]), \
                "AddressCell expected a cell '{}'".format(address)

            col_idx, row = (a or 0 for a in address[:2])
            column = (col_idx or '') and get_column_letter(col_idx)
            coordinate = '{0}{1}'.format(column, row or '')

        if sheet:
            format_str = '{0}!{1}'
        else:
            format_str = '{1}'

        return super(AddressCell, cls).__new__(
            cls, format_str.format(sheet, coordinate),
            sheet, col_idx, row, coordinate)

    def __str__(self):
        return self.address

    @property
    def is_range(self):
        return False

    @property
    def size(self):
        return AddressSize(1, 1)

    @property
    def has_sheet(self):
        return bool(self.sheet)

    @property
    def sort_key(self):
        return self.sheet, self.col_idx, self.row

    @property
    def column(self):
        return (self.col_idx or '') and get_column_letter(self.col_idx)

    def inc_col(self, inc):
        return (self.col_idx + inc - 1) % MAX_COL + 1

    def inc_row(self, inc):
        return (self.row + inc - 1) % MAX_ROW + 1

    def address_at_offset(self, row_inc=0, col_inc=0):
        new_col = self.inc_col(col_inc)
        new_row = self.inc_row(row_inc)
        return AddressCell((new_col, new_row, new_col, new_row),
                           sheet=self.sheet)

    @classmethod
    def create(cls, address, sheet='', cell=None):
        addr = AddressRange.create(address, sheet=sheet, cell=cell)
        if not isinstance(addr, AddressCell):
            raise ValueError(
                "{0} is not a valid coordinate".format(address))
        return addr


AddressSize = collections.namedtuple('AddressSize', 'height width')


def unquote_sheetname(sheetname):
    """
    Remove quotes from around, and embedded "''" in, quoted sheetnames

    sheetnames with special characters are quoted in formulas
    This is the inverse of openpyxl.utils.quote_sheetname
    """
    if sheetname.startswith("'") and sheetname.endswith("'"):
        sheetname = sheetname[1:-1].replace("''", "'")
    return sheetname


def split_sheetname(address, sheet=''):
    sh = ''
    if '!' in address:
        sh, address_part = address.split('!', maxsplit=1)
        assert '!' not in address_part, \
            "Only rectangular formulas are supported {}".format(address)
        sh = unquote_sheetname(sh)
        address = address_part

        if sh and sheet and sh != sheet:
            raise ValueError("Mismatched sheets '{}' and '{}'".format(
                sh, sheet))

    return sheet or sh, address


def structured_reference_boundaries(address, cell=None, sheet=None):
    # Excel reference: https://support.office.com/en-us/article/
    #   Using-structured-references-with-Excel-tables-
    #   F5ED2452-2337-4F71-BED3-C8AE6D2B276E

    match = TABLE_REF_RE.match(address)
    if not match:
        return None

    if cell is None:
        raise PyCelException(
            "Must pass cell for Structured Reference {}".format(address))

    name = match.group('table_name')
    table, sheet = cell.excel.table(name, sheet)
    table, sheet = cell.excel.table(name, sheet)

    if table is None:
        x = cell.excel.table(name, sheet)
        raise PyCelException(
            "Table {} not found for Structured Reference: {}".format(
                name, address))

    boundaries = openpyxl_range_boundaries(table.ref)
    assert None not in boundaries

    selector = match.group('table_selector')

    if not selector:
        # all columns and the data rows
        rows, start_col, end_col = None, None, None

    else:
        selector_match = TABLE_SELECTOR_RE.match(selector)
        if selector_match is None:
            raise PyCelException(
                "Unknown Structured Reference Selector: {}".format(selector))

        row_or_column = selector_match.group('row_or_column')
        this_row_column = selector_match.group('this_row_column')

        if row_or_column:
            rows = start_col = None
            end_col = row_or_column

        elif this_row_column:
            rows = '#This Row'
            start_col = None
            end_col = this_row_column

        else:
            rows = selector_match.group('rows')
            start_col = selector_match.group('start_col')
            end_col = selector_match.group('end_col')

            if rows is not None:
                if not rows:
                    rows = None

                elif '[' in rows:
                    rows = [r.split(']')[0] for r in rows.split('[')[1:]]
                    if len(rows) != 1:
                        # not currently supporting multiple row selects
                        raise PyCelException(
                            "Unknown Structured Reference Rows: {}".format(address))

                    rows = rows[0]

        if end_col.startswith('#'):
            # end_col collects the single field case
            if rows is None and start_col is None:
                rows = end_col
                end_col = None

        elif end_col.startswith('@'):
            rows = '#This Row'
            end_col = end_col[1:]
            if len(end_col) == 0:
                end_col = start_col

    if rows is None:
        # skip the headers and footers
        min_row = boundaries[1] + (table.headerRowCount if table.headerRowCount else 0)
        max_row = boundaries[3] - (table.totalsRowCount if table.totalsRowCount else 0)

    else:
        if rows == '#All':
            min_row, max_row = boundaries[1], boundaries[3]

        elif rows == '#Data':
            min_row = boundaries[1] + (table.headerRowCount if table.headerRowCount else 0)
            max_row = boundaries[3] - (table.totalsRowCount if table.totalsRowCount else 0)

        elif rows == '#Headers':
            min_row = boundaries[1]
            max_row = boundaries[1] + (table.headerRowCount if table.headerRowCount else 0) - 1

        elif rows == '#Totals':
            min_row = boundaries[3] - (table.totalsRowCount if table.totalsRowCount else 0) + 1
            max_row = boundaries[3]

        elif rows == '#This Row':
            # ::TODO:: If not in a data row, return #VALUE! How to do this?
            min_row = max_row = cell.address.row

        else:
            raise PyCelException(
                "Unknown Structured Reference Rows: {}".format(rows))

    if end_col is None:
        # all columns
        min_col_idx, max_col_idx = boundaries[0], boundaries[2]

    else:
        # a specific column
        column_idx = next((idx for idx, c in enumerate(table.tableColumns)
                           if c.name == end_col), None)
        if column_idx is None:
            raise PyCelException(
                "Column {} not found for Structured Reference: {}".format(
                    end_col, address))
        max_col_idx = boundaries[0] + column_idx

        if start_col is None:
            min_col_idx = max_col_idx

        else:
            column_idx = next((idx for idx, c in enumerate(table.tableColumns)
                               if c.name == start_col), None)
            if column_idx is None:
                raise PyCelException(
                    "Column {} not found for Structured Reference: {}".format(
                        start_col, address))
            min_col_idx = boundaries[0] + column_idx

    if min_row > max_row or min_col_idx > max_col_idx:
        raise PyCelException("Columns out of order : {}".format(address))

    return (min_col_idx, min_row, max_col_idx, max_row), sheet


def range_boundaries(address, cell=None, sheet=None):
    """
    R1C1 reference style

    You can also use a reference style where both the rows and the columns on
    the worksheet are numbered. The R1C1 reference style is useful for
    computing row and column positions in macros. In the R1C1 style, Excel
    indicates the location of a cell with an "R" followed by a row number
    and a "C" followed by a column number.

    Reference   Meaning

    R[-2]C      A relative reference to the cell two rows up and in
                the same column

    R[2]C[2]    A relative reference to the cell two rows down and
                two columns to the right

    R2C2        An absolute reference to the cell in the second row and
                in the second column

    R[-1]       A relative reference to the entire row above the active cell

    R           An absolute reference to the current row as part of a range

    """
    try:
        # if this is normal reference then just use the openpyxl converter
        boundaries = openpyxl_range_boundaries(address)
        if None not in boundaries or ':' in address:
            return boundaries, sheet
    except ValueError:
        pass

    m = R1C1_RANGE_RE.match(address)
    if not m:
        # Try to see if the is a structured table reference
        table_select_boundaries = structured_reference_boundaries(
            address, cell=cell, sheet=sheet)
        if table_select_boundaries:
            return table_select_boundaries

        # Try to see if this is a defined name
        name_addr = (cell and cell.excel and
                     cell.excel.defined_names.get(address))
        if name_addr:
            return openpyxl_range_boundaries(name_addr[0]), name_addr[1]

        raise ValueError(
            "{0} is not a valid coordinate or range".format(address))

    def from_relative_to_absolute(r1_or_c1):
        def require_cell():
            assert cell is not None, \
                "Must pass a cell to decode a relative address {}".format(
                    address)

        if not r1_or_c1.endswith(']'):
            if len(r1_or_c1) > 1:
                return int(r1_or_c1[1:])

            else:
                require_cell()
                if r1_or_c1[0].upper() == 'R':
                    return cell.row
                else:
                    return cell.col_idx

        else:
            require_cell()
            if r1_or_c1[0].lower() == 'r':
                return (cell.row + int(r1_or_c1[2:-1]) - 1) % MAX_ROW + 1
            else:
                return (cell.col_idx + int(r1_or_c1[2:-1]) - 1) % MAX_COL + 1

    min_col, min_row, max_col, max_row = (
        g if g is None else from_relative_to_absolute(g) for g in (
            m.group(n) for n in ('min_col', 'min_row', 'max_col', 'max_row')
        )
    )

    items_present = (min_col is not None, min_row is not None,
                     max_col is not None, max_row is not None)

    is_range = ':' in address
    if (is_range and items_present not in VALID_R1C1_RANGE_ITEM_COMBOS or
            not is_range and sum(items_present) < 2):
        raise ValueError(
            "{0} is not a valid coordinate or range".format(address))

    if min_col is not None:
        min_col = min_col

    if min_row is not None:
        min_row = min_row

    if max_col is not None:
        max_col = max_col
    else:
        max_col = min_col

    if max_row is not None:
        max_row = max_row
    else:
        max_row = min_row

    return (min_col, min_row, max_col, max_row), sheet


def resolve_range(address):
    """Return a list or nested lists with AddressCell for each element"""

    # ::TODO:: look at removing the assert
    assert isinstance(address, (AddressRange, AddressCell))

    # single cell, no range
    if not address.is_range:
        data = [address]

    else:

        start = address.start
        end = address.end

        # single column
        if start.column == end.column:
            data = list(next(address.cols))

        # single row
        elif start.row == end.row:
            data = list(next(address.rows))

        # rectangular range
        else:
            data = list(list(row) for row in address.rows)

    return data


def get_linest_degree(cell):
    # TODO: assumes a row or column of linest formulas &
    # that all coefficients are needed

    address = cell.address
    # figure out where we are in the row

    # to the left
    i = 0
    while True:
        i -= 1
        f = cell.excel.get_formula_from_range(
            address.address_at_offset(row_inc=0, col_inc=i))
        if not f or f != cell.formula:
            break

    # to the right
    j = 0
    while True:
        j += 1
        f = cell.excel.get_formula_from_range(
            address.address_at_offset(row_inc=0, col_inc=j))
        if not f or f != cell.formula:
            break

    # assume the degree is the number of linest's
    # last -1 is because an n degree polynomial has n+1 coefs
    degree = (j - i - 1) - 1

    # which coef are we (left most coef is the coef for the highest power)
    coef = -i

    # no linests left or right, try looking up/down
    if degree == 0:
        # up
        i = 0
        while True:
            i -= 1
            f = cell.excel.get_formula_from_range(
                address.address_at_offset(row_inc=i, col_inc=0))
            if not f or f != cell.formula:
                break

        # down
        j = 0
        while True:
            j += 1
            f = cell.excel.get_formula_from_range(
                address.address_at_offset(row_inc=j, col_inc=0))
            if not f or f != cell.formula:
                break

        degree = (j - i - 1) - 1
        coef = -i

    # if degree is zero -> only one linest formula
    # linear regression -> degree should be one
    return max(degree, 1), coef


def flatten(items):
    for item in items:
        if isinstance(item, collections.Iterable) and not isinstance(item, str):
            yield from flatten(item)
        else:
            yield item


def uniqueify(seq):
    seen = set()
    return tuple(x for x in seq if x not in seen and not seen.add(x))


def is_number(s):
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def coerce_to_number(value):
    if not isinstance(value, str):
        if is_number(value) and int(value) == float(value):
            return int(value)
        return value

    try:
        if value == DIV0:
            return 1 / 0
        elif '.' not in value:
            return int(value)
    except (ValueError, TypeError):
        pass

    try:
        return float(value)
    except (ValueError, TypeError):
        return value


def is_leap_year(year):
    if not is_number(year):
        raise TypeError("%s must be a number" % str(year))
    if year <= 0:
        raise TypeError("%s must be strictly positive" % str(year))

    # Watch out, 1900 is a leap according to Excel =>
    # https://support.microsoft.com/en-us/kb/214326
    return year % 4 == 0 and year % 100 != 0 or year % 400 == 0 or year == 1900


def get_max_days_in_month(month, year):
    if month == 2 and is_leap_year(year):
        return 29

    return calendar.monthrange(year, month)[1]


def normalize_year(y, m, d):
    """taking into account negative month and day values"""
    if m <= 0:
        y -= int(abs(m) / 12 + 1)
        m = 12 - (abs(m) % 12)
        normalize_year(y, m, d)
    elif m > 12:
        y += int(m / 12)
        m = m % 12

    if d <= 0:
        d += get_max_days_in_month(m, y)
        m -= 1
        y, m, d = normalize_year(y, m, d)

    else:
        days_in_month = get_max_days_in_month(m, y)
        if d > days_in_month:
            m += 1
            d -= days_in_month
            y, m, d = normalize_year(y, m, d)

    return y, m, d


def date_from_int(datestamp):

    if datestamp == 31 + 29:
        # excel thinks 1900 is a leap year
        return 1900, 2, 29

    date = dt.datetime(1899, 12, 30) + dt.timedelta(days=datestamp)
    if datestamp < 31 + 29:
        date += dt.timedelta(days=1)

    return date.year, date.month, date.day


def criteria_parser(criteria):
    if is_number(criteria):
        # numeric equals comparision
        def check(x):
            return is_number(x) and x == float(criteria)

    elif type(criteria) == str:

        search = re.search(r'(\W*)(.*)', criteria).group
        criteria_operator = search(1)
        op = OPERATORS[criteria_operator]
        value = search(2)

        # all operators except == (blank) are numeric
        numeric_compare = bool(criteria_operator) or is_number(value)

        def validate_number(x):
            if is_number(x):
                return True
            else:
                if numeric_compare:
                    raise TypeError(
                        'excellib.countif() doesnt\'t work for checking'
                        ' non number items against non equality')
                return False

        value = float(value) if validate_number(value) else str(value).lower()

        def check(x):
            if is_number(x):
                return op(x, value)
            else:
                return x.lower() == value

    else:
        raise ValueError("Couldn't parse criteria: {}".format(criteria))

    return check


def find_corresponding_index(rng, criteria):
    """This does not parse all of the patterns available to countif, etc"""
    # parse criteria
    if not isinstance(rng, list):
        raise TypeError('%s must be a list' % str(rng))

    if isinstance(criteria, list):
        # ugly...
        return ()

    # build a criteria check
    check = criteria_parser(criteria)

    return tuple(index for index, item in enumerate(rng) if check(item))


def build_operator_operand_fixup(capture_error_state):

    def fixup(left_op, op, right_op):
        """Fix up python operations to be more excel like in these cases:

            divide by zero
            value errors

            Empty cells
            Case-insensitive string compare
            String to Number coercion
            String / Number multiplication
        """

        if DIV0 in (left_op, right_op):
            return DIV0

        if VALUE_ERROR in (left_op, right_op):
            return VALUE_ERROR

        if left_op in (None, '#EMPTY!'):
            left_op = 0 if (not isinstance(
                right_op, str) or right_op == '#EMPTY!') else ''

        if right_op in (None, '#EMPTY!'):
            right_op = 0 if (not isinstance(
                left_op, str) or left_op == '#EMPTY!') else ''

        if op in ('Eq', 'NotEq'):
            if isinstance(left_op, str) and isinstance(right_op, str):
                left_op = left_op.lower()
                right_op = right_op.lower()

        elif op == 'BitAnd':
            # use bitwise-and '&' as string concat not '+'
            left_op = str(coerce_to_number(left_op))
            right_op = str(coerce_to_number(right_op))
            op = 'Add'

        else:
            left_op = coerce_to_number(left_op)
            right_op = coerce_to_number(right_op)

        if op == 'Mult':
            if isinstance(left_op, str) or isinstance(right_op, str):
                # Python is quite happy to multiply strings and numbers
                capture_error_state(
                    False,
                    "Cannot multiple type: {}({}) * {}({})".format(
                        type(left_op).__name__, left_op,
                        type(right_op).__name__, right_op
                    )
                )
                return VALUE_ERROR

        try:
            return PYTHON_AST_OPERATORS[op](left_op, right_op)
        except ZeroDivisionError:
            capture_error_state(
                True, 'Values: {} {} {}'.format(left_op, op, right_op))
            return DIV0
        except TypeError:
            capture_error_state(
                True, 'Values: {} {} {}'.format(left_op, op, right_op))
            return VALUE_ERROR

    return fixup
